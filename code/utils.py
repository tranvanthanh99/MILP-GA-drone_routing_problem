from matplotlib import markers
import matplotlib.pyplot as plt
import matplotlib as mpl
import math
import pandas as pd
import random


data = [
    {
        "node": 0,
        "coordinate": (0, 0),
        "label": "D",
        "demand": 0
    },
    {
        "node": 1,
        "coordinate": (0, 15),
        "label": "P",
        "demand": 2
    },
    {
        "node": 2,
        "coordinate": (7, 20),
        "label": "P",
        "demand": 5
    },
    {
        "node": 3,
        "coordinate": (25, 20),
        "label": "P",
        "demand": 5
    },
    {
        "node": 4,
        "coordinate": (25, 39),
        "label": "P",
        "demand": 3
    },
    {
        "node": 5,
        "coordinate": (14, 41),
        "label": "P",
        "demand": 1
    },
    {
        "node": 6,
        "coordinate": (17, 9),
        "label": "P",
        "demand": 3
    },
    {
        "node": 7,
        "coordinate": (13, 21),
        "label": "P",
        "demand": 2
    },
    {
        "node": 8,
        "coordinate": (4, 27),
        "label": "P",
        "demand": 2
    },
    {
        "node": 9,
        "coordinate": (26, 0),
        "label": "P",
        "demand": 2
    },
]

tour = [0, 1, 2, 8, 0, 3, 4, 5, 7, 0, 6, 9, 0]


def print_solution(data, tour, title="", save = False):
    
    #draw arc
    fig, ax = plt.subplots()
    for i in range(len(tour) - 1):
        # x= (i['src']['lon'], i['src']['lat'])
        # y = (i['des']['lon'], i['des']['lat'])
        x= (data[tour[i]]['coordinate'][0], data[tour[i]]['coordinate'][1])
        y= (data[tour[i+1]]['coordinate'][0], data[tour[i+1]]['coordinate'][1])

        size = 1000
        radius = math.sqrt(size)/2.
        arrow = mpl.patches.FancyArrowPatch(posA=x, posB=y, 
                                            arrowstyle='-|>', mutation_scale=20, 
                                            shrinkA=radius, shrinkB=radius)
        ax.add_patch(arrow)

    # draw label
    for idx, node in enumerate(data):
        # print(idx, node)
        demand = node['demand']
        plt.plot(node['coordinate'][0], node['coordinate'][1], marker='o', mfc='1', mec='g', markersize = 30)
        plt.plot(node['coordinate'][0], node['coordinate'][1], marker=f'${idx}, {demand}$', markersize = 20)

    # draw title
    plt.title(title)
    if save == True:
        plt.savefig("out.png")
    
    plt.show()

def to_excel(filename, data_ws, df, start_row=2, start_col=2):
    """Replacement for pandas .to_excel(). 

    For .xlsx and .xls formats only.

    args:
        start_row: df row +2; does not include header and is 1 based indexed.
    """
    writer = pd.ExcelWriter(filename.lower(), engine='openpyxl')
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    if data_ws not in wb.sheetnames:
        wb.create_sheet(data_ws)

    # Create the worksheet if it does not yet exist.
    writer.book = wb
    writer.sheets = {x.title: x for x in wb.worksheets}

    ws = writer.sheets[data_ws]
    # Fill with blanks.
    try:
        for row in ws:
            for cell in row:
                cell.value = None
    except TypeError:
        pass

    # Write manually to avoid overwriting formats.

    # Column names.
    ws.cell(1, 1).value = df.columns.name
    for icol, col_name in zip(range(2, len(df.columns) + 2), df.columns):
        ws.cell(1, icol).value = col_name

    # Row headers.
    for irow, row_name in zip(range(2, len(df.index) + 2), df.index):
        ws.cell(irow, 1).value = row_name

    # Body cells.
    for row, irow in zip([x[1] for x in df.iloc[start_row - 2:].iterrows()], list(range(start_row, len(df.index) + 2))):
        for cell, icol in zip([x[1] for x in row.iloc[start_col - 2:].items()], list(range(start_col, len(df.columns) + 2))):
            ws.cell(irow, icol).value = cell  # Skip the index.

    for row in ws.values:
        print('\t'.join(str(x or '') for x in row))
    print('Saving.')
    while True:
        try:
            writer.save()
            break
        except PermissionError:
            print(f'Please close {filename} before we can write to it!')
            time.sleep(2)
    writer.close()
    print('Done saving df.')

def gen_nodes(quantity=10):
    nodes=[
        [0, 0, 0],      # depot
        [0, 15, 28],    # recharging station
    ]
    index=['depot', 'RS']
    Pos_arr=[(0,0), (15,28)]
    # generate customer nodes
    for i in range(quantity):
        demand = random.randint(1, 5)
        PosX, PosY = 0, 0
        while (PosX, PosY) in Pos_arr:
            PosX, PosY = random.randint(0, 50), random.randint(0, 50)
        Pos_arr.append((PosX, PosY))
        nodes.append([demand, PosX, PosY])
        index.append(f'node{i+1}')
    df = pd.DataFrame(nodes, index=index, columns=['demand', 'PosX', 'PosY'])
    # df.to_excel('data.xlsx', sheet_name='nodes')
    to_excel('data.xlsx', 'nodes', df)
    

def read_nodes(quantity=0):
    df = pd.read_excel('data.xlsx', sheet_name='nodes')
    if quantity == 0 or quantity > len(df):
        quantity = len(df)
    else:
        quantity += 2
    nodes=[]
    row_index = df.index.values.tolist()
    for i in range(quantity):
        node={}
        node['label'] = row_index[i]
        node['demand'] = df['demand'][i]
        node['coordinate'] = (df['PosX'][i], df['PosY'][i])
        nodes.append(node)
    return nodes

def save_solution(num_of_nodes, cost, run_time):
    data=[]
    xl = pd.ExcelFile('data.xlsx')
    if 'GA' in xl.sheet_names:
        df = pd.read_excel('data.xlsx', sheet_name='GA')
        print(df['num_of_nodes'])
        for i in range(len(df['num_of_nodes'])):
            data.append([
                df['num_of_nodes'][i], df['cost'][i], df['run_time'][i]
            ])
    data.append([num_of_nodes, cost, run_time])
    df = pd.DataFrame(data, columns=['num_of_nodes', 'cost', 'run_time'])
    to_excel('data.xlsx', 'GA', df)

def plot_solution():
    import numpy as np
    df = pd.read_excel('data.xlsx', sheet_name='GA')
    df1 = pd.read_excel('data.xlsx', sheet_name='MILP')

    data1 = []
    data2 = []
    data3 = []

    for i in range(len(df1['cost'])):
        data3.append(df1['cost'][i])

    tol = 0
    index_arr1 = df.loc[df['num_of_nodes'] == 10].index.values.tolist()
    min_cost = df['cost'][index_arr1[0]]
    for i in index_arr1:
        tol += df['cost'][i]
        min_cost = min(min_cost, df['cost'][i])
    data1.append(tol / len(df.loc[df['num_of_nodes'] == 10]))
    data2.append(min_cost)

    tol = 0
    index_arr2 = df.loc[df['num_of_nodes'] == 15].index.values.tolist()
    min_cost = df['cost'][index_arr2[0]]
    for i in index_arr2:
        tol += df['cost'][i]
        min_cost = min(min_cost, df['cost'][i])
    data1.append(tol / len(df.loc[df['num_of_nodes'] == 15]))
    data2.append(min_cost)

    tol = 0
    index_arr3 = df.loc[df['num_of_nodes'] == 20].index.values.tolist()
    min_cost = df['cost'][index_arr3[0]]
    for i in index_arr3:
        tol += df['cost'][i]
        min_cost = min(min_cost, df['cost'][i])
    data1.append(tol / len(df.loc[df['num_of_nodes'] == 20]))
    data2.append(min_cost)

    width = 0.3
    labels=['Customer = 10', 'Customer = 15', 'Customer = 20']
    plt.bar(np.arange(len(data1)), data1, width=width)
    plt.bar(np.arange(len(data2)) + width, data2, width=width)
    plt.bar(np.arange(len(data2)) + width * 2, data3, width=width)
    plt.legend(labels=['GA overal cost', 'GA best cost', 'MILP'])
    plt.xticks([0 + width, 1 + width, 2 + width], labels)
    plt.ylabel('Cost')
    plt.show()

if __name__ == "__main__":
    # print_solution(data, tour)
    # gen_nodes(100)
    # print(read_nodes(10))
    # save_solution(10, 420, 0.6)
    plot_solution()
    pass